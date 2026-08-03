import datetime
import os
import shutil
import sys
from typing import List, Tuple

from models.database import _execute, _fetchone, get_connection, project_root, bundled_resource_path
from openpyxl import Workbook, load_workbook


TEMPLATE_COLUMNS = ['number', 'hieroglyph', 'pinyin', 'translation', 'phrase']


def template_path() -> str:
    return os.path.join(project_root(), 'templates', 'word_list_template.xlsx')


def ensure_template() -> str:
    path = template_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if os.path.exists(path):
        return path
    bundled = bundled_resource_path(os.path.join('templates', 'word_list_template.xlsx'))
    if os.path.exists(bundled):
        shutil.copy2(bundled, path)
        return path
    wb = Workbook()
    ws = wb.active
    ws.title = 'words'
    ws.append(TEMPLATE_COLUMNS)
    ws.append([1, '电脑', 'diànnǎo', 'компьютер', '我有一台电脑。'])
    wb.save(path)
    return path


def import_excel(user_id: int, file_path: str, list_name: str) -> Tuple[bool, str, int]:
    list_name = (list_name or os.path.splitext(os.path.basename(file_path))[0])[:15]
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        return False, f'Не удалось открыть файл: {exc}', 0

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return False, 'Файл пуст.', 0

    header = [str(c).strip().lower() if c is not None else '' for c in rows[0]]
    expected = [c.lower() for c in TEMPLATE_COLUMNS]
    if header[: len(expected)] != expected:
        return False, 'Неверный формат. Используйте шаблон Excel.', 0

    words: List[tuple] = []
    for idx, row in enumerate(rows[1:], start=1):
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        values = list(row[:5])
        while len(values) < 5:
            values.append('')
        number = values[0]
        try:
            number = int(number)
        except (TypeError, ValueError):
            number = idx
        words.append(
            (
                number,
                str(values[1] or '').strip(),
                str(values[2] or '').strip(),
                str(values[3] or '').strip(),
                str(values[4] or '').strip(),
            )
        )

    if not words:
        return False, 'В файле нет данных.', 0

    now = datetime.datetime.now().isoformat(timespec='seconds')
    conn = get_connection()
    cursor = conn.execute(
        'INSERT INTO custom_word_lists(user_id, name, card_count, created_at) VALUES(?,?,?,?)',
        (user_id, list_name, len(words), now),
    )
    list_id = cursor.lastrowid
    for word in words:
        conn.execute(
            'INSERT INTO custom_words(list_id, number, hieroglyph, pinyin, translation, phrase) '
            'VALUES(?,?,?,?,?,?)',
            (list_id, *word),
        )
    conn.commit()
    return True, list_name, list_id


def list_custom_lists(user_id: int) -> list:
    from models.database import _fetchall
    return _fetchall(
        'SELECT * FROM custom_word_lists WHERE user_id = ? ORDER BY list_id DESC',
        (user_id,),
    )


def list_all_custom_lists() -> list:
    from models.database import _fetchall
    return _fetchall(
        'SELECT custom_word_lists.*, users.user_name '
        'FROM custom_word_lists JOIN users ON users.user_id = custom_word_lists.user_id '
        'ORDER BY list_id DESC'
    )


def delete_custom_list(list_id: int) -> None:
    _execute('DELETE FROM custom_words WHERE list_id = ?', (list_id,))
    _execute('DELETE FROM custom_word_lists WHERE list_id = ?', (list_id,))
